import time
import requests
import re
from bs4 import BeautifulSoup
import openpyxl

# 新建表格
# wb = openpyxl.Workbook()
# ws = wb.active
# 继续之前的表格
excelBookPath = 'text.xlsx'
wb = openpyxl.load_workbook(excelBookPath)

wsAll = wb.create_sheet('总书本')
ws_1_1 = wb.create_sheet('玄幻小说')
ws_2_1 = wb.create_sheet('修真小说')
ws_3_1 = wb.create_sheet('都市小说')
ws_4_1 = wb.create_sheet('历史小说')
ws_5_1 = wb.create_sheet('网游小说')
ws_6_1 = wb.create_sheet('科幻小说')
ws_quanben = wb.create_sheet('完本小说')
ws_top = wb.create_sheet('更新日期')

# 表总行数
max_row = wsAll.max_row
# 表总列数
max_col = wsAll.max_column

wsAll['A1'] = '书ID'
wsAll['B1'] = '书名'
wsAll['C1'] = '作者名'
wsAll['D1'] = '作者页面'
wsAll['E1'] = '小说类型'
wsAll['F1'] = '类型url'
wsAll['G1'] = '书封面'


data_excel = []
header1={'user-agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36 SE 2.X MetaSr 1.0'}
#
header2 = {
    'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'
}
url = 'https://www.bbiquge.net/book'
url1 = 'https://www.bbiquge.net/book/{0}'


def Re(str, pat):
    patten = re.compile(pat)
    return patten.findall(str.replace(" ", ""))

if __name__ == '__main__':
    for j in range(0, 1200):
        wb = openpyxl.load_workbook(excelBookPath)
        wsAll = wb['总书本']
        booklists = []
        for i in range(100*j+1, 100*j+100):
            nowUlr = url1.format(i)
            try:
                if i % 2 == 0:
                    req = requests.get(url=nowUlr, headers=header1, timeout=10)
                else:
                    req = requests.get(url=nowUlr, headers=header2, timeout=10)
            except:
                print('接口请求有问题了，可能是被封了哦！{0}', format(i))
            try:
                soup = BeautifulSoup(req.text, features='html.parser')
                imglists = soup.find('div', class_='img_in')
                bookImg = BeautifulSoup(str(imglists), features='html.parser').find('img')['src']
                infolists = soup.find('div', id='info')
                nameSoup = BeautifulSoup(str(infolists), features='html.parser').find('h1')
                bookname = Re(nameSoup.get_text(), '(.+)/')
                bookId = str(i)
                makeBookUser = Re(nameSoup.get_text(), '/(.+)')
                aUrl = BeautifulSoup(str(nameSoup), features='html.parser').find('a')['href']
                bookUserUrl = url + aUrl
                target = soup.find('div', class_='fr').next_sibling.next_sibling.next_sibling.text
                targetUrl = url + soup.find('div', class_='fr').next_sibling.next_sibling.next_sibling['href']
                booklists.append([bookId, bookname[0], makeBookUser[0], bookUserUrl, target, targetUrl, bookImg])
                print(i)
                time.sleep(2)
            except:
                print("Id为{}的url 没有数据哦！", format(i))
                time.sleep(5)
        for i in booklists:
            wsAll.append(i)
        print("-{0} 数据已经爬取完了哦~--", format(j))
        wb.save('text.xlsx')
    print("----，------")
    print("-数据已经爬取完了哦~--")
    print("----，------")
