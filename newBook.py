import time
import requests
from bs4 import BeautifulSoup
import openpyxl
header1={'user-agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36 SE 2.X MetaSr 1.0'}
#
header2 = {
    'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'
}
url = 'https://www.bbiquge.net/fenlei/{0}_{1}/'


# 14898   ->  22999

wb = openpyxl.Workbook()

ws = wb.create_sheet('最近更新')
ws_1_1 = wb.create_sheet('玄幻小说')
ws_2_1 = wb.create_sheet('修真小说')
ws_3_1 = wb.create_sheet('都市小说')
ws_4_1 = wb.create_sheet('历史小说')
ws_5_1 = wb.create_sheet('网游小说')
ws_6_1 = wb.create_sheet('科幻小说')

ws['A1'] = '书名'
ws['B1'] = '书链接'
ws['C1'] = '文章名'
ws['D1'] = '文章链接'
ws['E1'] = '作者民'
ws['F1'] = '更新时间'


if __name__ == '__main__':
    for j in range(1, 6):
        try:
            for i in range(1, 600):
                excelBookPath = 'newBook.xlsx'
                wb = openpyxl.load_workbook(excelBookPath)
                ws = wb['最近更新']
                new_data = []
                nowUrl = url.format(j, i)
                if i % 2 == 0:
                    req = requests.get(url=nowUrl, headers=header1, timeout=10)
                else:
                    req = requests.get(url=nowUrl, headers=header2, timeout=10)
                ulSoup = BeautifulSoup(req.text, features='html.parser').find('ul', class_='titlelist').find_all('li')
                for item in ulSoup:
                    nameUrl = item.find('a', class_='name')['href']
                    aUrl = item.find('div', class_='zz').find('a')['href']
                    title = item.find('div', class_='zz').get_text()
                    userName = item.find('div', class_='author').get_text()
                    newTime = item.find('div', class_='sj').get_text()
                    new_data.append([nameUrl, aUrl, title, userName, newTime])
                print(i)
                for i in new_data:
                    ws.append(i)
            print('{0}以已经爬取成功！', format(j))
            wb.save('newBook.xlsx')
            time.sleep(5)
        except:
            print('--报错啦---')
    print('pycharm')