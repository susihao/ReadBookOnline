import os

import openpyxl

# 模糊匹配 搜索
from fuzzywuzzy import fuzz

allBooks = openpyxl.load_workbook('text.xlsx')
allBookSheets = allBooks['总书本']

bookUrl = 'https://www.bbiquge.net/book/{0}'

def toBook(id, bookname):
    os.system('python downloadBook.py', id, bookname)
    print('正在跳转下载页面哦！')

if __name__ == '__main__':
    for i in range(1, allBookSheets.max_row):
        if fuzz.token_set_ratio('末日', str(allBookSheets.cell(row=i, column=2).value)) > 40:
            print(allBookSheets.cell(row=i, column=1).value, allBookSheets.cell(row=i, column=2).value)


    print('小苏已经查完了哦！')


